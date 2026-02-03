#!/usr/bin/env python3
"""
Генерация XLSX-файла для импорта материалов в Google Sheets.
Колонки соответствуют парсеру MaterialsSheetsParserService (build_column_mapping_from_headers).
Лист: Materials (по умолчанию в API).
Запуск: pip install openpyxl && python tests/generate_materials_import_xlsx.py
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("Установите openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Куда сохранить (в корне проекта или в tests)
OUTPUT_DIR = Path(__file__).resolve().parent.parent
OUTPUT_FILE = OUTPUT_DIR / "materials_import_template.xlsx"

# Заголовки — имена колонок, распознаваемые парсером (course_uid, external_uid, title, type, url, ...)
HEADERS = [
    "course_uid",
    "external_uid",
    "title",
    "type",
    "url",
    "description",
    "caption",
    "order_position",
    "is_active",
]

# Примеры строк на основе реальных course_uid из БД (COURSE-PY-01, COURSE-MATH-01)
ROWS = [
    # course_uid, external_uid, title, type, url, description, caption, order_position, is_active
    (
        "COURSE-PY-01",
        "MAT-PY-01-INTRO",
        "Введение в Python",
        "link",
        "https://docs.python.org/3/tutorial/introduction.html",
        "Официальный туториал Python",
        "Ссылка на раздел Introduction",
        1,
        "true",
    ),
    (
        "COURSE-PY-01",
        "MAT-PY-01-VIDEO",
        "Видео: установка Python",
        "video",
        "https://www.youtube.com/watch?v=dQw4w9WgXcQ",
        "Как установить Python на Windows и macOS",
        "",
        2,
        "true",
    ),
    (
        "COURSE-PY-01",
        "MAT-PY-01-TEXT",
        "Краткая шпаргалка по типам",
        "text",
        "",
        "Текст материала: int, str, list, dict — основные типы в Python.",
        "",
        3,
        "true",
    ),
    (
        "COURSE-MATH-01",
        "MAT-MATH-01-LINK",
        "Линейная алгебра: основы",
        "link",
        "https://en.wikipedia.org/wiki/Linear_algebra",
        "Wikipedia — Linear algebra",
        "",
        1,
        "true",
    ),
    (
        "COURSE-MATH-01",
        "MAT-MATH-01-PDF",
        "Конспект по матрицам",
        "pdf",
        "https://example.com/matrices.pdf",
        "PDF-документ с формулами",
        "",
        2,
        "true",
    ),
    (
        "COURSE-PY-01",
        "MAT-PY-01-SCRIPT",
        "Скрипт hello.py",
        "script",
        "https://example.com/hello.py",
        "Пример Python-скрипта",
        "",
        4,
        "true",
    ),
    (
        "COURSE-MATH-01",
        "MAT-MATH-01-DOC",
        "Лекция в DOCX",
        "document",
        "https://example.com/lecture.docx",
        "Документ с лекцией",
        "",
        3,
        "true",
    ),
]


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materials"

    # Заголовки
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)

    # Данные
    for row_idx, row_tuple in enumerate(ROWS, start=2):
        for col_idx, value in enumerate(row_tuple, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Создан файл: {OUTPUT_FILE}")
    print("Колонки:", ", ".join(HEADERS))
    print("Строк данных:", len(ROWS))
    print("Лист: Materials — укажите его в API как sheet_name или оставьте по умолчанию.")


if __name__ == "__main__":
    main()
