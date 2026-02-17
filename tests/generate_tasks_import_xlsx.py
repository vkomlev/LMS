#!/usr/bin/env python3
"""
Генерация XLSX-файла для импорта заданий из Google Sheets.
Колонки соответствуют парсеру SheetsParserService (parse_task_row).
Лист: Задания (по умолчанию в API).
Запуск: pip install openpyxl && python tests/generate_tasks_import_xlsx.py
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("Установите openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Куда сохранить (в корне проекта или в tests)
OUTPUT_DIR = Path(__file__).resolve().parent.parent
OUTPUT_FILE = OUTPUT_DIR / "tasks_import_test.xlsx"

# Два реальных course_uid из БД (можно поменять под вашу БД)
COURSE_UID_MAIN = "PY"
COURSE_UID_SUB = "PY-pishem-pervuju-programmu-na-python-peremennye-i-konstanty"

# Заголовки — имена колонок, распознаваемые парсером
HEADERS = [
    "external_uid",
    "course_uid",
    "type",
    "stem",
    "options",
    "correct_answer",
    "max_score",
    "code",
    "title",
    "prompt",
    "input_link",
    "accepted_answers",
]

# Для части заданий задаём другой курс, чтобы проверить импорт "курс на строке"
ROW_COURSE_UID: dict[str, str] = {
    "TEST-SC-002": COURSE_UID_SUB,
    "TEST-MC-002": COURSE_UID_SUB,
    "TEST-SA-002": COURSE_UID_SUB,
    "TEST-SA-COM-001": COURSE_UID_SUB,
    "TEST-SA-COM-002": COURSE_UID_SUB,
    "TEST-TA-002": COURSE_UID_SUB,
}

# Тестовые данные заданий разных типов
# Формат (без course_uid — он будет добавлен при записи в XLSX):
# (external_uid, type, stem, options, correct_answer, max_score, code, title, prompt, input_link, accepted_answers)
ROWS = [
    # SC (Single Choice) - задания с одним правильным ответом
    (
        "TEST-SC-001",
        "SC",
        "Что такое переменная в Python?",
        "A: Именованная область памяти для хранения данных | B: Функция для вывода данных | C: Тип данных | D: Оператор",
        "A",
        "10",
        "PY-VAR-001",
        "Переменные Python",
        "Переменная хранит значение, которое можно изменять",
        "",
        "",
    ),
    (
        "TEST-SC-002",
        "SC",
        "Какой оператор используется для целочисленного деления в Python?",
        "A: / | B: // | C: % | D: **",
        "B",
        "10",
        "PY-OP-001",
        "Операторы деления",
        "",
        "",
        "",
    ),
    (
        "TEST-SC-003",
        "SC",
        "Что выведет код: print(type([]))?",
        "A: <class 'list'> | B: <class 'tuple'> | C: <class 'dict'> | D: <class 'set'>",
        "A",
        "15",
        "PY-TYPE-001",
        "Типы данных",
        "",
        "",
        "",
    ),
    
    # MC (Multiple Choice) - задания с несколькими правильными ответами
    (
        "TEST-MC-001",
        "MC",
        "Какие из перечисленных типов данных являются неизменяемыми (immutable) в Python?",
        "A: list | B: tuple | C: str | D: dict | E: int",
        "B,C,E",
        "15",
        "PY-IMMUT-001",
        "Неизменяемые типы",
        "Неизменяемые типы нельзя изменить после создания",
        "",
        "",
    ),
    (
        "TEST-MC-002",
        "MC",
        "Выберите все способы создания словаря в Python:",
        "A: dict() | B: {} | C: dict(key='value') | D: [key: value]",
        "A,B,C",
        "20",
        "PY-DICT-001",
        "Создание словарей",
        "",
        "",
        "",
    ),
    (
        "TEST-MC-003",
        "MC",
        "Какие из следующих операторов возвращают булево значение?",
        "A: == | B: = | C: != | D: in | E: and",
        "A,C,D",
        "15",
        "PY-BOOL-001",
        "Булевы операторы",
        "",
        "",
        "",
    ),
    
    # SA (Short Answer) - задания с коротким ответом
    (
        "TEST-SA-001",
        "SA",
        "Сколько байт занимает тип int в Python 3?",
        "",
        "8",
        "10",
        "PY-INT-001",
        "Размер int",
        "",
        "",
        "8 | восемь | 8 байт",
    ),
    (
        "TEST-SA-002",
        "SA",
        "Какой метод используется для добавления элемента в конец списка?",
        "",
        "append",
        "10",
        "PY-LIST-001",
        "Методы списков",
        "",
        "",
        "append | .append() | append()",
    ),
    (
        "TEST-SA-003",
        "SA",
        "Какое максимальное значение может хранить тип int в Python 3?",
        "",
        "нет ограничений",
        "15",
        "PY-INT-002",
        "Ограничения int",
        "",
        "",
        "нет ограничений | неограничен | unlimited | бесконечно",
    ),
    
    # SA_COM (Short Answer with Comments) - задания с коротким ответом и комментариями
    (
        "TEST-SA-COM-001",
        "SA_COM",
        "Напишите функцию для вычисления факториала числа n.",
        "",
        "def factorial(n):\n    if n <= 1:\n        return 1\n    return n * factorial(n-1)",
        "20",
        "PY-FUNC-001",
        "Функция факториала",
        "Используйте рекурсию или цикл",
        "",
        "",
    ),
    (
        "TEST-SA-COM-002",
        "SA_COM",
        "Что выведет следующий код?\n\nx = [1, 2, 3]\ny = x\ny.append(4)\nprint(x)",
        "",
        "[1, 2, 3, 4]",
        "15",
        "PY-LIST-002",
        "Мутация списков",
        "Обратите внимание на то, что y ссылается на тот же объект, что и x",
        "",
        "[1, 2, 3, 4] | [1,2,3,4]",
    ),
    
    # TA (Text Answer) - задания с развернутым ответом (требуют ручной проверки)
    (
        "TEST-TA-001",
        "TA",
        "Объясните разницу между списком (list) и кортежем (tuple) в Python. Приведите примеры использования каждого типа данных.",
        "",
        "Нет правильного ответа",
        "25",
        "PY-TYPES-001",
        "Сравнение типов",
        "Опишите основные различия и когда использовать каждый тип",
        "",
        "",
    ),
    (
        "TEST-TA-002",
        "TA",
        "Опишите принципы работы декораторов в Python. Напишите пример декоратора, который измеряет время выполнения функции.",
        "",
        "Нет правильного ответа",
        "30",
        "PY-DECOR-001",
        "Декораторы",
        "Включите объяснение синтаксиса и практический пример",
        "",
        "",
    ),
    (
        "TEST-TA-003",
        "TA",
        "Объясните концепцию генераторов (generators) в Python. В чем их преимущества перед обычными списками?",
        "",
        "Нет правильного ответа",
        "25",
        "PY-GEN-001",
        "Генераторы",
        "Опишите yield и преимущества в памяти",
        "",
        "",
    ),
    
    # Дополнительные задания для разнообразия
    (
        "TEST-SC-004",
        "SC",
        "Какой синтаксис используется для создания множества (set) в Python?",
        "A: set() | B: {} | C: [] | D: ()",
        "A",
        "10",
        "PY-SET-001",
        "Множества",
        "",
        "",
        "",
    ),
    (
        "TEST-MC-004",
        "MC",
        "Какие из следующих конструкций используются для обработки исключений в Python?",
        "A: try-except | B: try-finally | C: try-else | D: catch-throw",
        "A,B,C",
        "20",
        "PY-EXCEPT-001",
        "Обработка исключений",
        "",
        "",
        "",
    ),
    (
        "TEST-SA-004",
        "SA",
        "Как называется специальный метод, который вызывается при создании объекта класса?",
        "",
        "__init__",
        "10",
        "PY-CLASS-001",
        "Специальные методы",
        "",
        "",
        "__init__ | __init__() | init",
    ),
    (
        "TEST-SA-005",
        "SA",
        "Какой оператор используется для проверки принадлежности элемента последовательности?",
        "",
        "in",
        "10",
        "PY-OP-002",
        "Оператор in",
        "",
        "",
        "in",
    ),
    (
        "TEST-TA-004",
        "TA",
        "Опишите концепцию наследования классов в Python. Приведите пример базового класса и производного класса с переопределением метода.",
        "",
        "Нет правильного ответа",
        "30",
        "PY-INHERIT-001",
        "Наследование",
        "Включите пример кода с базовым и производным классом",
        "",
        "",
    ),
]


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Заголовки
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)

    # Данные
    for row_idx, row_tuple in enumerate(ROWS, start=2):
        values = row_tuple
        # Поддержка старого формата ROWS без course_uid
        if len(row_tuple) == len(HEADERS) - 1:
            external_uid = str(row_tuple[0])
            course_uid = ROW_COURSE_UID.get(external_uid, COURSE_UID_MAIN)
            values = (external_uid, course_uid, *row_tuple[1:])

        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Автоподбор ширины колонок
    for col in range(1, len(HEADERS) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col)
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
        ws.column_dimensions[column_letter].width = adjusted_width

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUTPUT_FILE)
        saved_path = OUTPUT_FILE
    except PermissionError:
        # Частый кейс на Windows: файл открыт в Excel и заблокирован
        alt_file = OUTPUT_DIR / "tasks_import_test__new.xlsx"
        wb.save(alt_file)
        saved_path = alt_file

    print(f"[OK] Создан файл: {saved_path}")
    print(f"[INFO] Колонки: {', '.join(HEADERS)}")
    print(f"[INFO] Строк данных: {len(ROWS)}")
    print(f"[INFO] Лист: 'Задания' - укажите его в API как sheet_name или оставьте по умолчанию.")
    print("\n[INFO] Типы заданий в файле:")
    print(f"   - SC (Single Choice): {sum(1 for r in ROWS if r[1] == 'SC')}")
    print(f"   - MC (Multiple Choice): {sum(1 for r in ROWS if r[1] == 'MC')}")
    print(f"   - SA (Short Answer): {sum(1 for r in ROWS if r[1] == 'SA')}")
    print(f"   - SA_COM (Short Answer with Comments): {sum(1 for r in ROWS if r[1] == 'SA_COM')}")
    print(f"   - TA (Text Answer): {sum(1 for r in ROWS if r[1] == 'TA')}")
    print("\n[INFO] Для импорта используйте:")
    print("   POST /api/v1/tasks/import/google-sheets")
    print("   с параметрами: course_code='PY', difficulty_code='NORMAL'")


if __name__ == "__main__":
    main()
